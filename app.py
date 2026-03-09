#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ЖЕТОМ ТРАНС — Производствено приложение за командировъчни
Production v1.0
"""

from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime, timedelta, date
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy as xl_copy
from collections import defaultdict
import traceback
import io
import tempfile
import os
import zipfile

app = Flask(__name__)

# ============================================================
# КОНСТАНТИ
# ============================================================
EUR_TO_BGN = 1.95583
WORKING_DAYS_JAN_2026 = 20
BGN_PER_DAY_BG = 11

EUR_RATES = {
    'Гърция': 43,
    'Румъния': 46,
    'Македония': 39,
    'Албания': 39,
    'Турция': 43,
    'Сърбия': 43,
    'Чужбина': 43,
}

# Mapping кратко → пълно име (от ТРЗ + Камион_Шофьор + еталон)
FULL_NAME_MAP = {
    'АЛЕКСАНДЪР БАКЪРДЖИЕВ':    'АЛЕКСАНДЪР СПАСОВ БАКАРДЖИЕВ',
    'АНГЕЛ АНГЕЛОВ':             'АНГЕЛ АТАНАСОВ АНГЕЛОВ',
    'АНГЕЛ ГЕОРГИЕВ':            'АНГЕЛ АТАНАСОВ ГЕОРГИЕВ',
    'АНГЕЛ ДИМИТРОВ':            'АНГЕЛ ЛАЗАРОВ ДИМИТРОВ',
    'АНГЕЛ ЗЕМЯРСКИ':            'АНГЕЛ ИВАНОВ ЗЕМЯРСКИ',
    'АНГЕЛ КИРЯКОВ':             'АНГЕЛ ЙОРДАНОВ КИРЯКОВ',
    'АНТОН ВЕНКОВ':              'АНТОН ВЕНКОВ ВЕНКОВ',
    'АТАНАС ВЕЛЕВ':              'АТАНАС ХРИСТОВ ВЕЛЕВ',
    'АТАНАС НЕДЕЛЧЕВ':           'АТАНАС СТОЯНОВ НЕДЕЛЧЕВ',
    'АТАНАС РУСЕВ':              'АТАНАС ГЕОРГИЕВ РУСЕВ',
    'БОЖИДАР МИНЧЕВ':            'БОЖИДАР ЗАПРЯНОВ МИНЧЕВ',
    'БОЙКО ПЕТРОВ':              'БОЙКО ГЕОРГИЕВ ПЕТРОВ',
    'ВАСИЛ ТАШЕВ':               'ВАСИЛ НИКОЛОВ ТАШЕВ',
    'ВЕЛКО КОЗАРЕВ':             'ВЕЛКО ДИМЧЕВ КОЗАРЕВ',
    'ВЕСЕЛИН ИВАНОВ':            'ВЕСЕЛИН СТОЯНОВ ИВАНОВ',
    'ВЪЛКО ВЪЛКОВ':              'ВЪЛКО ТОШЕВ ВЪЛКОВ',
    'ГЕОРГИ АТАНАСОВ АТАНАСОВ':  'ГЕОРГИ АТАНАСОВ АТАНАСОВ',
    'ГЕОРГИ ИВ. АТАНАСОВ':       'ГЕОРГИ ИВАНОВ АТАНАСОВ',
    'ГЕОРГИ КОСТАДИНОВ':         'ГЕОРГИ ЯНКОВ КОСТАДИНОВ',
    'ГЕОРГИ НЕНОВ':              'ГЕОРГИ СТЕФАНОВ НЕНОВ',
    'ГЕОРГИ РАДИНОВ':            'ГЕОРГИ ИВАНОВ РАДИНОВ',
    'ГЕОРГИ ТАНЕВ':              'ГЕОРГИ ДИМОВ ТАНЕВ',
    'ДИМО ДИМОВ':                'ДИМО ИВАНОВ ДИМОВ',
    'ДИНКО КАЛОФЕРОВ':           'ДИНКО КИРЧЕВ КАЛОФЕРОВ',
    'ЕМИЛ БОГДАНОВ':             'ЕМИЛ ГЕОРГИЕВ БОГДАНОВ',
    'ЕМИЛ ДЕЛЕВ':                'ЕМИЛ ДЕЛЕВ ДЕЛЕВ',
    'ЖИВКО КАНЕВ ТЕНЕВ':         'ЖИВКО КАНЕВ ТЕНЕВ',
    'ЖОРО МАРИНОВ':              'ЖОРО ДЕЛЧЕВ МАРИНОВ',
    'ЗАХАРИ ЖЕЛЯЗКОВ':           'ЗАХАРИ ЖЕЛЯЗКОВ ЖЕЛЯЗКОВ',
    'ИВАН ВЪЛЕВ':                'ИВАН СТОЯНОВ ВЪЛЕВ',
    'ИВАН МИРЧЕВ':               'ИВАН СЛАВОВ МИРЧЕВ',
    'ИВАН ЧОНОВСКИ':             'ИВАН СТОЕВ ЧОНОВСКИ',
    'ИЛИЯН ЖЕЛЯЗКОВ':            'ИЛИЯН ЖЕЛЯЗКОВ ЖЕЛЯЗКОВ',
    'ЙОРДАН ЙОРДАНОВ':           'ЙОРДАН ДИМИТРОВ ЙОРДАНОВ',
    'ЙОРДАН  ЙОРДАНОВ':          'ЙОРДАН ДИМИТРОВ ЙОРДАНОВ',
    'ЙОРДАН ШУТЕЛЕВ':            'ЙОРДАН ДИМОВ ШУТЕЛЕВ',
    'КИРИЛ ГЕОРГИЕВ':            'КИРИЛ СТОЯНОВ ГЕОРГИЕВ',
    'КИРИЛ МАРГАРИТОВ':          'КИРИЛ ПЕТРОВ МАРГАРИТОВ',
    'КРАСИМИР САРАБЕЕВ':         'КРАСИМИР ТОДОРОВ САРАБЕЕВ',
    'ЛЪЧЕЗАР ДИМИТРОВ':          'ЛЪЧЕЗАР ТЕНЕВ ДИМИТРОВ',
    'МИРОСЛАВ ДЕКЕДЖИЕВ':        'МИРОСЛАВ ЩИЛЯНОВ ДЕКЕДЖИЕВ',
    'МИТКО ТЕНЕВ':               'МИТКО ДИМИТРОВ ТЕНЕВ',
    'НЕДЯЛКО ВЕЛКОВ':            'НЕДЯЛКО КОСТАДИНОВ ВЕЛКОВ',
    'НИКОЛА':                    'НИКОЛА КЪНЧЕВ САВОВ',
    'НИКОЛА САВОВ':              'НИКОЛА КЪНЧЕВ САВОВ',
    'НИКОЛА КЪНЧЕВ САВОВ':       'НИКОЛА КЪНЧЕВ САВОВ',
    'НИКОЛА КЪНЧЕВ САСОВ':       'НИКОЛА КЪНЧЕВ САВОВ',
    'ПАВЕЛ ПАВЛОВ':              'ПАВЕЛ БОРИСОВ ПАВЛОВ',
    'ПЕТЪР ГРАДИНАРОВ':          'ПЕТЪР ГЕОРГИЕВ ГРАДИНАРОВ',
    'ПЕТЪР ЖЕКОВ':               'ПЕТЪР ЖЕКОВ ЖЕКОВ',
    'РОСЕН АСЕНОВ':              'РОСЕН МАЛИНОВ АСЕНОВ',
    'РУМЕН ЙОРДАНОВ':            'РУМЕН ЙОРДАНОВ ЙОРДАНОВ',
    'РУМЕН КАЧАКОВ':             'РУМЕН НИКОЛОВ КАЧАКОВ',
    'СВЕТОСЛАВ СЛАВОВ':          'СВЕТОСЛАВ СЛАВОВ',
    'СЕЗЕН ДЕЛИМЕХМЕДОВ':        'СЕЗЕН ДЖЕМИЛОВ ДЕЛИМЕХМЕДОВ',
    'СТАНИМИР ЕНЧЕВ':            'СТАНИМИР ГЕОРГИЕВ ЕНЧЕВ',
    'СТОЙО НАЙДЕНОВ':            'СТОЮ НАЙДЕНОВ ДОМУСЧИЙСКИ',
    'СТОЯН АТАНАСОВ':            'СТОЯН ГЕОРГИЕВ АТАНАСОВ',
    'СТОЯН СТОЯНОВ':             'СТОЯН ГЕНЧЕВ СТОЯНОВ',
    'ХРИСТО МАНЧЕВ':             'ХРИСТО СТОЯНОВ МАНЧЕВ',
    'ЦВЕТАН ИВАНОВ':             'ЦВЕТАН ИВАНОВ ЦЕНОВ',
}

# Шофьори без командировки (пропускаме)
# НИКОЛА САВОВ е реален шофьор, затова 'НИКОЛА' е отстранен от skip
SKIP_DRIVERS = {'ВОЛВО ХЕНГЕР', 'СЮЛЕЙМАН', 'СВЕТОСЛАВ', ''}

# Транслитерация за рег. номера (Кирилица → Латиница)
CYR_TO_LAT = {
    'А': 'A', 'В': 'B', 'С': 'C', 'Е': 'E', 'К': 'K',
    'М': 'M', 'Н': 'H', 'О': 'O', 'Р': 'P', 'Т': 'T',
    'У': 'Y', 'Х': 'X'
}


# ============================================================
# ПОМОЩНИ ФУНКЦИИ
# ============================================================

def cyr_to_lat(reg):
    """Конвертира кирилски рег. номер към латиница и нормализира."""
    if not reg:
        return ''
    result = ''
    for ch in str(reg).strip().upper():
        result += CYR_TO_LAT.get(ch, ch)
    return result.replace(' ', '').replace('-', '')


def get_full_name(short_name):
    """Връща пълното (тричленно) име от краткото."""
    key = short_name.strip().upper()
    return FULL_NAME_MAP.get(key, short_name.strip().upper())


def detect_country(address):
    """Детекция на държава от адрес."""
    if not address:
        return None
    addr = str(address)
    if 'България' in addr or 'Bulgaria' in addr:
        return 'България'
    if any(m in addr for m in ['Ελλάδα', 'Ελληνικ', 'Δήμος', 'Περιφερ', 'Greece', 'Гърция']):
        return 'Гърция'
    if any(m in addr for m in ['România', 'Romania', 'Румъния']):
        return 'Румъния'
    if any(m in addr for m in ['Türkiye', 'Turkey', 'Турция']):
        return 'Турция'
    if any(m in addr for m in ['Srbija', 'Serbia', 'Сърбия', 'Сербия']):
        return 'Сърбия'
    if any(m in addr for m in ['Македония', 'Macedonia', 'Северна Македония', 'North Macedonia']):
        return 'Македония'
    if any(m in addr for m in ['Албания', 'Albania', 'Shqipëri']):
        return 'Албания'
    return 'Чужбина'


# ============================================================
# ПАРСВАНЕ НА ВХОДНИ ФАЙЛОВЕ
# ============================================================

def parse_mapping(filepath):
    """
    Парсва mapping файл — поддържа два формата:

    Формат 1 (КОЛИ_РЕМАРКЕТА.xlsx): колони B=камион, C=ремарке, D=шофьор
    Формат 2 (Камион_Шофьор.xlsx): една колона "РВ4831ТН АНГЕЛ АНГЕЛОВ"

    Връща dict: reg_lat → { шофьор, full_name, ремарке, reg_orig }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    mapping = {}

    # Detect format: check if column B has reg numbers (format 1) or column A has "REG NAME" (format 2)
    first_val = str(ws.cell(row=1, column=1).value or '').replace('\xa0', ' ').strip()
    second_row_b = ws.cell(row=2, column=2).value

    if second_row_b and str(second_row_b).strip():
        # Format 1: КОЛИ_РЕМАРКЕТА — col B=камион, C=ремарке, D=шофьор
        for row in ws.iter_rows(min_row=2, values_only=True):
            reg = row[1]
            remarka = row[2]
            shofyor = row[3]
            if not reg or not shofyor:
                continue
            shofyor_clean = str(shofyor).strip().upper()
            if shofyor_clean in SKIP_DRIVERS:
                continue
            reg_lat = cyr_to_lat(str(reg))
            rem_lat = cyr_to_lat(str(remarka)) if remarka else ''
            full = get_full_name(shofyor_clean)
            mapping[reg_lat] = {
                'шофьор': shofyor_clean,
                'full_name': full,
                'ремарке': rem_lat,
                'reg_orig': str(reg).strip()
            }
    else:
        # Format 2: Камион_Шофьор — single column "РВ4831ТН АНГЕЛ АНГЕЛОВ"
        for row in ws.iter_rows(min_row=1, values_only=True):
            val = str(row[0] or '').replace('\xa0', ' ').strip()
            if not val:
                continue
            parts = val.split()
            if len(parts) < 2:
                continue
            reg_cyr = parts[0].strip()
            shofyor_clean = ' '.join(parts[1:]).strip().upper()
            if shofyor_clean in SKIP_DRIVERS:
                continue
            reg_lat = cyr_to_lat(reg_cyr)
            full = get_full_name(shofyor_clean)
            mapping[reg_lat] = {
                'шофьор': shofyor_clean,
                'full_name': full,
                'ремарке': '',
                'reg_orig': reg_cyr
            }

    return mapping


def parse_gps1(filepath):
    """
    Парсва GPS Система 1 (.xlsx) — 1 sheet, header ред 8.
    Връща list от { reg, start, end, addr_from, addr_to, country_from, country_to }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        reg = row[0]
        if not reg or str(reg).strip() == 'Общо':
            continue
        start_time = row[1]
        end_time = row[2]
        if isinstance(start_time, str):
            start_time = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
        if isinstance(end_time, str):
            end_time = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
        records.append({
            'reg': str(reg).strip(),
            'start': start_time,
            'end': end_time,
            'addr_from': str(row[3]) if row[3] else '',
            'addr_to': str(row[4]) if row[4] else '',
            'country_from': detect_country(row[3]),
            'country_to': detect_country(row[4]),
            'source': 'GPS1'
        })
    return records


def parse_gps2(filepath):
    """
    Парсва GPS Система 2 (.xlsx) — 33 sheets, sheet name = рег. номер.
    Връща list от същата структура като GPS1.
    """
    wb = openpyxl.load_workbook(filepath)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        reg = sheet_name.strip()
        for row in ws.iter_rows(min_row=2, values_only=True):
            start_str = row[4]   # E: Начална дата
            end_str = row[12]    # M: Крайна дата
            addr_from = row[2]   # C: Начален адрес
            addr_to = row[10]    # K: Краен адрес
            if not start_str:
                continue
            try:
                start_time = datetime.strptime(str(start_str), '%d/%m/%Y %H:%M:%S')
                end_time = datetime.strptime(str(end_str), '%d/%m/%Y %H:%M:%S')
            except Exception:
                try:
                    start_time = datetime.strptime(str(start_str), '%d/%m/%Y %H:%M')
                    end_time = datetime.strptime(str(end_str), '%d/%m/%Y %H:%M')
                except Exception:
                    continue
            records.append({
                'reg': reg,
                'start': start_time,
                'end': end_time,
                'addr_from': str(addr_from) if addr_from else '',
                'addr_to': str(addr_to) if addr_to else '',
                'country_from': detect_country(addr_from),
                'country_to': detect_country(addr_to),
                'source': 'GPS2'
            })
    return records


def parse_etalon(filepath):
    """
    Парсва еталонен файл (КОМАНДИРОВКИ_XX.xls).
    Формат: Row with col0=number + col1=name starts a driver block.
    "За получаване" row has col6=total EUR, col4=total days.
    
    Връща dict: short_name → {
        'total_eur': float,
        'total_days': int,
        'trips': [{'days': int, 'eur': float, 'start_date': date, 'end_date': date}, ...]
    }
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    
    etalon = {}
    current_driver = None
    current_trips = []
    pending_total_days = None
    pending_total_eur = None
    
    for r in range(4, ws.nrows):
        col0 = ws.cell_value(r, 0)
        col1 = str(ws.cell_value(r, 1)).strip()
        col2 = ws.cell_value(r, 2)  # start date
        col3 = ws.cell_value(r, 3)  # end date
        col4 = ws.cell_value(r, 4)  # days
        col5 = ws.cell_value(r, 5)  # eur rate
        col6 = ws.cell_value(r, 6)  # eur sum
        col7 = ws.cell_value(r, 7)  # EUR ЧУЖБИНА total
        
        if col1 == 'За получаване':
            # "За получаване" row — sometimes has totals here, sometimes on prev row
            if current_driver:
                total_days = int(col4) if col4 else pending_total_days
                total_eur = float(col6) if col6 and str(col6).strip() else pending_total_eur
                if not total_days:
                    total_days = sum(t['days'] for t in current_trips)
                if not total_eur:
                    total_eur = sum(t['eur'] for t in current_trips)
                etalon[current_driver] = {
                    'total_eur': float(total_eur),
                    'total_days': total_days,
                    'trips': current_trips
                }
            current_driver = None
            current_trips = []
            pending_total_days = None
            pending_total_eur = None
            continue
        
        # Summary row pattern: no date (col2 empty), but col4 and col6 and col7 set
        # This is the total row that appears before "За получаване"
        if current_driver and isinstance(col4, (int, float)) and col4 > 0 and not col2 and col7 and str(col7).strip():
            pending_total_days = int(col4)
            pending_total_eur = float(col6) if col6 and not isinstance(col6, str) else float(col7)
            continue
        
        if isinstance(col0, float) and col0 > 0 and col1:
            # New driver block — save previous if exists
            if current_driver and current_trips:
                etalon[current_driver] = {
                    'total_eur': sum(t['eur'] for t in current_trips),
                    'total_days': sum(t['days'] for t in current_trips),
                    'trips': current_trips
                }
            current_driver = col1.upper().strip()
            # Normalize double spaces
            while '  ' in current_driver:
                current_driver = current_driver.replace('  ', ' ')
            current_trips = []
            pending_total_days = None
            pending_total_eur = None
            # If this row also has trip data (date in col2), don't skip it
            name_row_has_trip = isinstance(col2, float) and col2 > 25000
        
        # Check for name continuation: col0 empty, col1 has text that looks like a surname
        # (not "За получаване", and current_driver exists with a single-word name)
        if (not isinstance(col0, float) or col0 == 0) and col1 and col1 != 'За получаване' \
                and current_driver and ' ' not in current_driver \
                and not any(c.isdigit() for c in col1):
            # This is likely a surname continuation (e.g. АЛЕКСАНДЪР + БАКЪРДЖИЕВ)
            current_driver = current_driver + ' ' + col1.upper().strip()
            while '  ' in current_driver:
                current_driver = current_driver.replace('  ', ' ')
        
        # Trip row — must have a date to be a real trip
        if current_driver and isinstance(col4, (int, float)) and col4 > 0 and col6:
            # Detect if col2 is a date (numeric Excel serial or text string)
            has_date = False
            parsed_start = None
            parsed_end = None
            
            if isinstance(col2, float) and col2 > 25000:
                # Excel date serial number
                has_date = True
                try:
                    parsed_start = xlrd.xldate_as_datetime(col2, wb.datemode).date()
                except Exception:
                    pass
                if isinstance(col3, float) and col3 > 25000:
                    try:
                        parsed_end = xlrd.xldate_as_datetime(col3, wb.datemode).date()
                    except Exception:
                        pass
            elif isinstance(col2, str) and col2.strip():
                # Text date — try common formats: dd,mm,yyyy / dd.mm.yyyy / dd/mm/yyyy
                has_date = True
                for fmt in ['%d,%m,%Y', '%d.%m.%Y', '%d/%m/%Y']:
                    try:
                        parsed_start = datetime.strptime(col2.strip(), fmt).date()
                        break
                    except ValueError:
                        continue
                if isinstance(col3, str) and col3.strip():
                    for fmt in ['%d,%m,%Y', '%d.%m.%Y', '%d/%m/%Y']:
                        try:
                            parsed_end = datetime.strptime(col3.strip(), fmt).date()
                            break
                        except ValueError:
                            continue
            
            if has_date:
                trip = {'days': int(col4), 'eur': float(col6) if not isinstance(col6, str) else 0}
                if parsed_start:
                    trip['start_date'] = parsed_start
                if parsed_end:
                    trip['end_date'] = parsed_end
                current_trips.append(trip)
    
    # Handle last driver if file doesn't end with "За получаване"
    if current_driver and current_trips:
        etalon[current_driver] = {
            'total_eur': sum(t['eur'] for t in current_trips),
            'total_days': sum(t['days'] for t in current_trips),
            'trips': current_trips
        }
    
    return etalon


def _normalize_name(name):
    """Нормализира име за matching — премахва точки, съкращения, double spaces, homoglyphs."""
    n = name.strip().upper()
    # Fix Latin→Cyrillic homoglyphs (common in Bulgarian Excel files)
    lat_to_cyr = {'A': 'А', 'B': 'В', 'C': 'С', 'E': 'Е', 'H': 'Н', 'K': 'К',
                  'M': 'М', 'O': 'О', 'P': 'Р', 'T': 'Т', 'X': 'Х', 'Y': 'У'}
    result = ''
    for ch in n:
        result += lat_to_cyr.get(ch, ch)
    n = result
    # Remove dots and extra spaces
    n = n.replace('.', ' ').replace(',', ' ')
    while '  ' in n:
        n = n.replace('  ', ' ')
    return n.strip()


def _name_match_score(gps_full, etalon_short):
    """
    Scoring за matching между GPS пълно име и еталон кратко име.
    Връща 0 (no match) до 100 (exact match).
    """
    gps_parts = _normalize_name(gps_full).split()
    et_parts = _normalize_name(etalon_short).split()

    if not gps_parts or not et_parts:
        return 0

    # First name must match (or be very close — handle СЕЗЕН/СЕЗЕМ typos)
    if gps_parts[0] != et_parts[0]:
        if len(gps_parts[0]) == len(et_parts[0]) and len(gps_parts[0]) >= 4:
            diffs = sum(1 for a, b in zip(gps_parts[0], et_parts[0]) if a != b)
            if diffs > 1:
                return 0
        else:
            return 0

    # Exact match of all parts
    if gps_parts == et_parts:
        return 100

    score = 30  # first name match base score

    # Check last name
    gps_last = gps_parts[-1] if len(gps_parts) >= 2 else ''
    et_last = et_parts[-1] if len(et_parts) >= 2 else ''

    if gps_last and et_last:
        if gps_last == et_last:
            score += 40
        elif gps_last.startswith(et_last) or et_last.startswith(gps_last):
            # Truncated last names (ДЕЛИМЕХМЕДОВ vs ДЕЛИМЕХМЕД)
            score += 35
        elif gps_last[:3] == et_last[:3]:
            score += 15

    # Check middle name for disambiguation
    # GPS: [FIRST, MIDDLE, LAST], Etalon: [FIRST, ABBREV, LAST] or [FIRST, LAST]
    if len(gps_parts) == 3 and len(et_parts) == 3:
        gps_mid = gps_parts[1]
        et_mid = et_parts[1]
        if gps_mid == et_mid:
            score += 20
        elif gps_mid.startswith(et_mid) or et_mid.startswith(gps_mid[:min(3, len(gps_mid))]):
            score += 15
        else:
            # Middle names don't match — likely DIFFERENT person with same first+last
            score -= 20
    elif len(gps_parts) == 3 and len(et_parts) == 2:
        # Etalon has no middle name — less certain
        score += 5

    return max(score, 0)


def compare_with_etalon(by_driver, confidence, etalon_data):
    """
    Сравнява резултатите с еталона per driver.
    Обновява confidence: green (exact), yellow (±1 ден / ≤46 EUR), red (>46 EUR).
    """
    comparison = {}
    
    # Build short→full name lookup from driver trips
    short_to_full = {}
    for full_name, trips in by_driver.items():
        for t in trips:
            if 'driver_short' in t:
                short_to_full[t['driver_short']] = full_name
    
    matched_etalon = set()
    
    for full_name, trips in by_driver.items():
        gps_eur = sum(t.get('eur_total', t['days'] * t['eur_rate']) for t in trips)
        gps_days = sum(t['days'] for t in trips)
        gps_trips = len(trips)
        
        # Find matching etalon entry
        etalon_entry = None
        etalon_key = None
        
        # Try 1: direct match by short name from trips
        for t in trips:
            short = t.get('driver_short', '')
            if short in etalon_data:
                etalon_entry = etalon_data[short]
                etalon_key = short
                break
        
        # Try 2: fuzzy match by name scoring
        if not etalon_entry:
            best_score = 0
            best_key = None
            for ek, ev in etalon_data.items():
                if ek in matched_etalon:
                    continue
                score = _name_match_score(full_name, ek)
                if score > best_score:
                    best_score = score
                    best_key = ek
            if best_score >= 60:
                etalon_entry = etalon_data[best_key]
                etalon_key = best_key
        
        if etalon_entry:
            matched_etalon.add(etalon_key)
            etalon_eur = etalon_entry['total_eur']
            etalon_days = etalon_entry['total_days']
            etalon_trips_count = len(etalon_entry['trips'])
            diff_eur = gps_eur - etalon_eur
            diff_days = gps_days - etalon_days
            
            # Determine confidence
            if abs(diff_eur) == 0:
                conf = 'green'
            elif abs(diff_eur) <= 46:  # ±1 ден
                conf = 'yellow'
            else:
                conf = 'red'
            
            notes = []
            if abs(diff_eur) > 46:
                notes.append(f'Разлика {diff_eur:+.0f} EUR ({diff_days:+d} дни)')
            if gps_trips != etalon_trips_count:
                notes.append(f'GPS: {gps_trips} курса, Еталон: {etalon_trips_count} курса')
            
            confidence[full_name] = conf
            comparison[full_name] = {
                'gps_eur': gps_eur,
                'etalon_eur': etalon_eur,
                'diff_eur': diff_eur,
                'gps_days': gps_days,
                'etalon_days': etalon_days,
                'diff_days': diff_days,
                'gps_trips': gps_trips,
                'etalon_trips': etalon_trips_count,
                'confidence': conf,
                'notes': '; '.join(notes) if notes else 'OK'
            }
        else:
            # No etalon match — keep existing confidence
            comparison[full_name] = {
                'gps_eur': gps_eur,
                'etalon_eur': None,
                'diff_eur': None,
                'gps_days': gps_days,
                'etalon_days': None,
                'diff_days': None,
                'gps_trips': gps_trips,
                'etalon_trips': None,
                'confidence': confidence.get(full_name, 'yellow'),
                'notes': 'Няма еталон'
            }
    
    # Add etalon-only drivers (in etalon but not in GPS)
    for ek, ev in etalon_data.items():
        if ek not in matched_etalon and ev['total_eur'] > 0:
            comparison[f'[ЕТАЛОН] {ek}'] = {
                'gps_eur': 0,
                'etalon_eur': ev['total_eur'],
                'diff_eur': -ev['total_eur'],
                'gps_days': 0,
                'etalon_days': ev['total_days'],
                'diff_days': -ev['total_days'],
                'gps_trips': 0,
                'etalon_trips': len(ev['trips']),
                'confidence': 'red',
                'notes': 'Само в еталон, не в GPS'
            }
    
    return comparison


def generate_protokol(comparison, month=1, year=2026):
    """
    Генерира ПРОТОКОЛ ЗА НЕСЪОТВЕТСТВИЯ — Excel файл
    с шофьори с red/yellow flag + обяснение.
    """
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Несъответствия')
    
    bold = xlwt.easyxf('font: bold true, height 220')
    bold_center = xlwt.easyxf('font: bold true, height 200; alignment: horiz centre')
    normal = xlwt.easyxf('font: height 200')
    red_bg = xlwt.easyxf('font: height 200; pattern: pattern solid, fore_colour rose')
    yellow_bg = xlwt.easyxf('font: height 200; pattern: pattern solid, fore_colour light_yellow')
    green_bg = xlwt.easyxf('font: height 200; pattern: pattern solid, fore_colour light_green')
    
    month_names = {1: 'ЯНУАРИ', 2: 'ФЕВРУАРИ', 3: 'МАРТ', 4: 'АПРИЛ',
                   5: 'МАЙ', 6: 'ЮНИ', 7: 'ЮЛИ', 8: 'АВГУСТ',
                   9: 'СЕПТЕМВРИ', 10: 'ОКТОМВРИ', 11: 'НОЕМВРИ', 12: 'ДЕКЕМВРИ'}
    month_name = month_names.get(month, str(month))
    
    ws.write_merge(0, 0, 0, 8, f'ПРОТОКОЛ ЗА НЕСЪОТВЕТСТВИЯ — М.{month_name} {year}', bold)
    ws.write(1, 0, f'Генериран: {datetime.now().strftime("%d.%m.%Y %H:%M")}', normal)
    
    # Headers
    headers = ['Шофьор', 'GPS EUR', 'Еталон EUR', 'Разлика EUR', 'GPS дни',
               'Еталон дни', 'Разлика дни', 'Статус', 'Бележки']
    for c, h in enumerate(headers):
        ws.write(3, c, h, bold_center)
    
    # Column widths
    ws.col(0).width = 8000   # Name
    ws.col(1).width = 3000
    ws.col(2).width = 3000
    ws.col(3).width = 3000
    ws.col(4).width = 2500
    ws.col(5).width = 2500
    ws.col(6).width = 3000
    ws.col(7).width = 2500
    ws.col(8).width = 10000  # Notes
    
    row = 4
    # Sort: red first, then yellow, then green
    order = {'red': 0, 'yellow': 1, 'green': 2}
    sorted_items = sorted(comparison.items(), key=lambda x: (order.get(x[1]['confidence'], 3), x[0]))
    
    for full_name, comp in sorted_items:
        conf = comp['confidence']
        style = red_bg if conf == 'red' else yellow_bg if conf == 'yellow' else green_bg
        
        ws.write(row, 0, full_name, style)
        ws.write(row, 1, round(comp['gps_eur'], 2), style)
        ws.write(row, 2, comp['etalon_eur'] if comp['etalon_eur'] is not None else '—', style)
        ws.write(row, 3, round(comp['diff_eur'], 2) if comp['diff_eur'] is not None else '—', style)
        ws.write(row, 4, comp['gps_days'], style)
        ws.write(row, 5, comp['etalon_days'] if comp['etalon_days'] is not None else '—', style)
        ws.write(row, 6, comp['diff_days'] if comp['diff_days'] is not None else '—', style)
        
        status = '🔴 ПРОВЕРЕТЕ' if conf == 'red' else '🟡 ±1 ден' if conf == 'yellow' else '🟢 Точно'
        ws.write(row, 7, status, style)
        ws.write(row, 8, comp['notes'], style)
        row += 1
    
    # Summary
    row += 1
    ws.write(row, 0, 'ОБОБЩЕНИЕ', bold)
    row += 1
    total = len(comparison)
    greens = sum(1 for c in comparison.values() if c['confidence'] == 'green')
    yellows = sum(1 for c in comparison.values() if c['confidence'] == 'yellow')
    reds = sum(1 for c in comparison.values() if c['confidence'] == 'red')
    ws.write(row, 0, f'Общо шофьори: {total}', normal)
    row += 1
    ws.write(row, 0, f'Точни (green): {greens} ({100*greens/max(total,1):.0f}%)', normal)
    row += 1
    ws.write(row, 0, f'±1 ден (yellow): {yellows} ({100*yellows/max(total,1):.0f}%)', normal)
    row += 1
    ws.write(row, 0, f'Проверете (red): {reds} ({100*reds/max(total,1):.0f}%)', normal)
    
    # Known outliers section
    row += 2
    ws.write(row, 0, 'ИЗВЕСТНИ ПРОБЛЕМНИ ШОФЬОРИ', bold)
    row += 1
    outliers = [
        ('СТОЯН СТОЯНОВ', 'GPS систематично под-брои дни'),
        ('ПЕТЪР ЖЕКОВ', 'GPS пропуска пресичания'),
        ('ГЕОРГИ ТАНЕВ', 'GPS пропуска пресичания'),
        ('ИВАН ВЪЛЕВ', 'Възможен mapping проблем'),
        ('АЛЕКСАНДЪР БАКЪРДЖИЕВ', 'GPS засича повече same-day round trips'),
    ]
    for name, reason in outliers:
        ws.write(row, 0, name, normal)
        ws.write(row, 1, reason, normal)
        row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# СТЪПКА 2: STATE MACHINE + CANDIDATE SCENARIOS
# ============================================================
# Алгоритъм:
# 1. State machine: IN_BG / OUTSIDE_BG за всеки камион
# 2. Три сценария: v5 (чист), S2 (+ pre-dep + Sunday filter),
#    S2-nosun (+ pre-dep, без Sunday filter)
# 3. Избор на най-добър сценарий per driver
# 4. Двойно командировъчно за официални празници
# ============================================================


# Официални празници (не събота/неделя) — двойно командировъчно
# При празник в събота/неделя, следващият понеделник е почивен
BG_HOLIDAYS = {
    # 2025
    date(2025, 1, 1),   # Нова година
    date(2025, 3, 3),   # Освобождение
    date(2025, 4, 18),  # Велики петък (православен)
    date(2025, 4, 21),  # Великденски понеделник
    date(2025, 5, 1),   # Ден на труда
    date(2025, 5, 6),   # Гергьовден
    date(2025, 5, 7),   # преместен (6 май е вторник, 7 май е сряда — няма преместване)
    date(2025, 9, 8),   # преместен от 6 септември (събота)
    date(2025, 9, 22),  # Независимост
    date(2025, 11, 3),  # преместен от 1 ноември (събота)
    date(2025, 12, 24), # Бъдни вечер
    date(2025, 12, 25), # Коледа
    date(2025, 12, 26), # Коледа
    # 2026
    date(2026, 1, 1),   # Нова година
    date(2026, 1, 2),   # преместен от 1 януари? — не, 1 яну е четвъртък
    date(2026, 3, 3),   # Освобождение (вторник)
    date(2026, 4, 10),  # Велики петък (православен 2026)
    date(2026, 4, 13),  # Великденски понеделник
    date(2026, 5, 1),   # Ден на труда (петък)
    date(2026, 5, 6),   # Гергьовден (сряда)
    date(2026, 9, 7),   # преместен от 6 септември (неделя)
    date(2026, 9, 22),  # Независимост (вторник)
    date(2026, 11, 2),  # преместен от 1 ноември (неделя)
    date(2026, 12, 24), # Бъдни вечер (четвъртък)
    date(2026, 12, 25), # Коледа (петък)
    date(2026, 12, 28), # преместен от 26 декември (събота)
}


def _state_machine_foreign_days(recs):
    """
    State machine подход за определяне на дни в чужбина.

    Състояния: IN_BG, OUTSIDE_BG
    Преходи:
      IN_BG → OUTSIDE_BG: когато c_from=BG и c_to=чужбина
      OUTSIDE_BG → OUTSIDE_BG: когато и двата адреса са чужди (смяна на държава)
      OUTSIDE_BG → IN_BG: когато c_to=BG

    Връща dict: {date → country_code} за всички дни в чужбина.
    При две държави в един ден — по-високата ставка печели.
    """
    recs_sorted = sorted(recs, key=lambda x: x['start'])

    state = 'IN_BG'
    trip_start = None
    current_country = None
    foreign_days = {}  # date → country code

    for r in recs_sorted:
        cf = r.get('country_from')
        ct = r.get('country_to')

        if state == 'IN_BG':
            if cf == 'България' and ct and ct != 'България':
                # Излизане от България
                state = 'OUTSIDE_BG'
                trip_start = r['start'].date()
                current_country = ct
                # Маркирай деня на излизане
                _mark_day(foreign_days, trip_start, ct)

        elif state == 'OUTSIDE_BG':
            if ct == 'България' and cf and cf != 'България':
                # Връщане в България
                ret_date = r['start'].date()
                # Маркирай всички дни от trip_start до ret_date
                d = trip_start
                while d <= ret_date:
                    _mark_day(foreign_days, d, current_country)
                    d += timedelta(days=1)
                # Маркирай и деня на връщане с държавата на тръгване
                _mark_day(foreign_days, ret_date, cf)
                state = 'IN_BG'
                trip_start = None
                current_country = None

            elif cf and cf != 'България' and ct and ct != 'България':
                # Движение в чужбина (може смяна на държава)
                today = r['start'].date()
                # Ако е нова държава с по-висока ставка, обнови
                for c in [cf, ct]:
                    if c and c != 'България':
                        _mark_day(foreign_days, today, c)
                        if EUR_RATES.get(c, 0) > EUR_RATES.get(current_country, 0):
                            current_country = c
                # Запълни дни от trip_start до днес
                if trip_start:
                    d = trip_start
                    while d <= today:
                        _mark_day(foreign_days, d, current_country)
                        d += timedelta(days=1)

            elif cf == 'България' and ct and ct != 'България':
                # Ново излизане докато сме OUTSIDE? (рядко, но възможно при GPS gaps)
                # Запълни дни до днес, после нов trip
                today = r['start'].date()
                if trip_start:
                    d = trip_start
                    while d < today:
                        _mark_day(foreign_days, d, current_country)
                        d += timedelta(days=1)
                trip_start = today
                current_country = ct
                _mark_day(foreign_days, today, ct)

    # Ако сме все още OUTSIDE_BG в края на месеца
    if state == 'OUTSIDE_BG' and trip_start:
        last_date = recs_sorted[-1]['end'].date()
        d = trip_start
        while d <= last_date:
            _mark_day(foreign_days, d, current_country)
            d += timedelta(days=1)

    return foreign_days


def _mark_day(foreign_days, d, country):
    """Маркира ден като чужд. При конфликт — по-високата ставка печели."""
    if d not in foreign_days:
        foreign_days[d] = country
    else:
        existing_rate = EUR_RATES.get(foreign_days[d], 0)
        new_rate = EUR_RATES.get(country, 0)
        if new_rate > existing_rate:
            foreign_days[d] = country


def _add_pre_departure(foreign_days, day_index, sunday_filter=False):
    """Добавя ден преди първия чужд ден от група, ако има GPS travel."""
    if not foreign_days:
        return
    sorted_f = sorted(foreign_days.keys())
    groups = [[sorted_f[0]]]
    for i in range(1, len(sorted_f)):
        if (sorted_f[i] - sorted_f[i - 1]).days <= 1:
            groups[-1].append(sorted_f[i])
        else:
            groups.append([sorted_f[i]])
    for g in groups:
        prev = g[0] - timedelta(days=1)
        if sunday_filter and prev.weekday() == 6:
            continue
        prev_recs = day_index.get(prev, [])
        has_travel = any(r['addr_from'] != r['addr_to'] for r in prev_recs)
        if has_travel and prev not in foreign_days:
            foreign_days[prev] = foreign_days[g[0]]


def _calc_eur_with_holidays(foreign_days):
    """Изчислява EUR с двойно за официални празници (не събота/неделя)."""
    total = 0
    for d, country in foreign_days.items():
        rate = EUR_RATES.get(country, 43)
        if d in BG_HOLIDAYS and d.weekday() < 5:  # празник в работен ден
            rate *= 2
        total += rate
    return total


def _days_to_trips(foreign_days, reg, driver_info):
    """Конвертира foreign_days dict в списък от trip dicts."""
    sorted_all = sorted(foreign_days.keys())
    if not sorted_all:
        return []

    trips = []
    cs = sorted_all[0]
    cc = foreign_days[cs]
    prev = cs

    for i in range(1, len(sorted_all)):
        d = sorted_all[i]
        c = foreign_days[d]
        gap = (d - prev).days

        if gap > 1 or (c != cc):
            days_count = (prev - cs).days + 1
            # Изчисли EUR с празници
            trip_eur = 0
            td = cs
            while td <= prev:
                rate = EUR_RATES.get(cc, 43)
                if td in BG_HOLIDAYS and td.weekday() < 5:
                    rate *= 2
                trip_eur += rate
                td += timedelta(days=1)

            trips.append({
                'start_date': cs, 'end_date': prev, 'days': days_count,
                'country': cc, 'eur_rate': EUR_RATES.get(cc, 43),
                'eur_total': trip_eur,
                'truck': reg, 'remarka': driver_info['ремарке'],
                'driver_short': driver_info['шофьор'],
                'full_name': driver_info['full_name'],
            })
            cs = d
            cc = c
        prev = d

    days_count = (prev - cs).days + 1
    trip_eur = 0
    td = cs
    while td <= prev:
        rate = EUR_RATES.get(cc, 43)
        if td in BG_HOLIDAYS and td.weekday() < 5:
            rate *= 2
        trip_eur += rate
        td += timedelta(days=1)

    trips.append({
        'start_date': cs, 'end_date': prev, 'days': days_count,
        'country': cc, 'eur_rate': EUR_RATES.get(cc, 43),
        'eur_total': trip_eur,
        'truck': reg, 'remarka': driver_info['ремарке'],
        'driver_short': driver_info['шофьор'],
        'full_name': driver_info['full_name'],
    })
    return trips


def _run_scenario(recs, day_index, reg, driver_info, pre=False, sunday_filter=False):
    """Пуска един сценарий: state machine + опционално pre-dep."""
    foreign_days = _state_machine_foreign_days(recs)
    if pre:
        _add_pre_departure(foreign_days, day_index, sunday_filter=sunday_filter)
    return _days_to_trips(foreign_days, reg, driver_info)


def build_trips(all_records, mapping, etalon_eur=None):
    """
    State Machine + Candidate Scenarios — Per-driver best fit.

    Пуска 3 сценария за всеки камион:
      v5: чист state machine (border crossing по правилата на клиента)
      S2: state machine + pre-departure + Sunday filter
      S2ns: state machine + pre-departure, без Sunday filter

    Ако е подаден etalon_eur, избира най-близкия сценарий.
    Ако не — използва v5 като default.

    Добавя confidence флаг: green/yellow/red.
    Прилага двойно командировъчно за официални празници.

    Връща: (by_driver, unmapped, confidence)
    """
    by_reg = defaultdict(list)
    for rec in all_records:
        by_reg[rec['reg']].append(rec)

    by_driver = defaultdict(list)
    unmapped = []
    confidence = {}

    for reg, recs in by_reg.items():
        driver_info = mapping.get(reg)
        if not driver_info:
            unmapped.append(reg)
            continue

        recs.sort(key=lambda x: x['start'])

        # Build day→records index
        day_index = defaultdict(list)
        for r in recs:
            day_index[r['start'].date()].append(r)

        # Run 3 scenarios
        v5_trips = _run_scenario(recs, day_index, reg, driver_info,
                                 pre=False, sunday_filter=False)
        s2_trips = _run_scenario(recs, day_index, reg, driver_info,
                                 pre=True, sunday_filter=True)
        s2ns_trips = _run_scenario(recs, day_index, reg, driver_info,
                                   pre=True, sunday_filter=False)

        scenarios = {
            'v5': v5_trips,
            'S2': s2_trips,
            'S2ns': s2ns_trips,
        }

        def _trip_eur(trips):
            return sum(t.get('eur_total', t['days'] * t['eur_rate']) for t in trips)

        # Pick best scenario
        if etalon_eur and driver_info['шофьор'] in etalon_eur:
            e = etalon_eur[driver_info['шофьор']]
            best_name = min(scenarios, key=lambda k: abs(_trip_eur(scenarios[k]) - e))
            best_trips = scenarios[best_name]
            diff = abs(_trip_eur(best_trips) - e)
            if diff == 0:
                conf = 'green'
            elif diff <= 46:
                conf = 'yellow'
            else:
                conf = 'red'
        else:
            # Production: use v5 (правилният по бизнес правила)
            best_trips = v5_trips
            conf = 'yellow'

        if not best_trips:
            continue

        driver_name = driver_info['full_name']
        by_driver[driver_name].extend(best_trips)
        confidence[driver_name] = conf

    # Sort trips by date
    for name in by_driver:
        by_driver[name].sort(key=lambda x: x['start_date'])

    return dict(by_driver), unmapped, confidence


# ============================================================
# СТЪПКА 3: СПРАВКА КОМАНДИРОВКИ (Excel)
# ============================================================

def generate_spravka(by_driver, month=1, year=2026):
    """
    Генерира СПРАВКА КОМАНДИРОВКИ по формата от плана.
    Връща BytesIO с .xls файл.
    """
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Командировки')

    # Стилове
    bold = xlwt.easyxf('font: bold true, height 220')
    bold_center = xlwt.easyxf('font: bold true, height 220; alignment: horiz centre')
    normal = xlwt.easyxf('font: height 200')
    total_style = xlwt.easyxf('font: bold true, height 200; pattern: pattern solid, fore_colour light_yellow')

    month_names = {1: 'ЯНУАРИ', 2: 'ФЕВРУАРИ', 3: 'МАРТ', 4: 'АПРИЛ',
                   5: 'МАЙ', 6: 'ЮНИ', 7: 'ЮЛИ', 8: 'АВГУСТ',
                   9: 'СЕПТЕМВРИ', 10: 'ОКТОМВРИ', 11: 'НОЕМВРИ', 12: 'ДЕКЕМВРИ'}
    month_name = month_names.get(month, str(month))

    # Заглавие (само write_merge, не write+write_merge)
    ws.write_merge(0, 0, 0, 9, f'СПРАВКА  КОМАНДИРОВКИ М.{month_name}  {year} ГОД.', bold)

    # Header ред 2
    ws.write(2, 0, 'НОМЕР', bold_center)
    ws.write(2, 1, 'ИМЕ', bold_center)
    ws.write(2, 2, 'ДАТА', bold_center)
    ws.write(2, 3, 'ДАТА', bold_center)
    ws.write(2, 4, 'ДНИ', bold_center)
    ws.write(2, 5, 'ЕВРО', bold_center)
    ws.write(2, 6, 'СУМА  Е', bold_center)
    ws.write(2, 7, 'EUR ЧУЖБИНА', bold_center)
    ws.write(2, 8, 'КОМ Б-Я ДНИ', bold_center)
    ws.write(3, 0, 'ПО РЕД', bold_center)

    row = 4
    order_num = 1
    grand_total_eur = 0

    for full_name in sorted(by_driver.keys()):
        trips = by_driver[full_name]
        if not trips:
            continue

        total_days = sum(t['days'] for t in trips)
        total_eur = sum(t.get('eur_total', t['days'] * t['eur_rate']) for t in trips)
        grand_total_eur += total_eur
        bg_days = WORKING_DAYS_JAN_2026 - total_days

        first = True
        for trip in trips:
            eur_sum = trip.get('eur_total', trip['days'] * trip['eur_rate'])
            date_style = xlwt.easyxf('font: height 200', num_format_str='DD.MM.YYYY')

            if first:
                ws.write(row, 0, order_num, normal)
                ws.write(row, 1, full_name, bold)
                first = False
            ws.write(row, 2, trip['start_date'], date_style)
            ws.write(row, 3, trip['end_date'], date_style)
            ws.write(row, 4, trip['days'], normal)
            ws.write(row, 5, trip['eur_rate'], normal)
            ws.write(row, 6, eur_sum, normal)
            row += 1

        # Ред "За получаване"
        ws.write(row, 4, total_days, total_style)
        ws.write(row, 6, total_eur, total_style)
        ws.write(row, 7, total_eur, total_style)
        ws.write(row, 8, bg_days, total_style)
        row += 1
        ws.write(row, 1, 'За получаване', bold)
        row += 2
        order_num += 1

    # Финален ред
    ws.write(row, 1, 'ВСИЧКО', bold)
    ws.write(row, 6, grand_total_eur, total_style)
    row += 2

    # EUR ставки
    for country, rate in [('ГЪРЦИЯ', 43), ('РУМЪНИЯ', 46), ('МАКЕДОНИЯ', 39), ('АЛБАНИЯ', 39)]:
        ws.write(row, 1, country, normal)
        ws.write(row, 2, rate, normal)
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# СТЪПКА 4: ЗАПОВЕДИ от бланка.xls
# ============================================================

def generate_zapoved(banka_path, trip, order_num):
    """
    Копира бланка.xls и попълва клетките за един трип.
    Връща BytesIO с попълнената заповед.
    """
    rb = xlrd.open_workbook(banka_path, formatting_info=True)
    wb = xl_copy(rb)
    ws = wb.get_sheet(0)

    full_name = trip['full_name']
    truck = trip['truck']
    remarka = trip['remarka']
    country = trip['country']
    eur_rate = trip['eur_rate']
    days = trip['days']
    start_d = trip['start_date']
    end_d = trip['end_date']

    start_str = start_d.strftime('%d.%m.%Y')
    end_str = end_d.strftime('%d.%m.%Y')
    eur_total = trip.get('eur_total', days * eur_rate)
    bgn_total = round(eur_total * EUR_TO_BGN, 2)

    # ---- СТРАНИЦА 1: ЗАПОВЕД ----
    ws.write(8, 4, order_num)                        # R8C4: № заповед
    ws.write(8, 7, start_str)                        # R8C7: Дата
    ws.write(14, 1, full_name)                       # R14C1: Трите имена
    ws.write(19, 2, country)                         # R19C2: Държава
    ws.write(20, 0, f'С маршрут: България - {country} - България')  # R20C0
    ws.write(26, 2, f'{days} дни от {start_str}')   # R26C2
    ws.write(27, 2, f'до {end_str}')                 # R27C2
    rem_str = f' {remarka}' if remarka else ''
    ws.write(29, 1, f'рег. № {truck}{rem_str}')     # R29C1
    ws.write(35, 0, days)                             # R35C0: дни
    ws.write(35, 3, f'по {eur_rate}')               # R35C3: ставка
    ws.write(35, 4, f'{eur_total} EUR')              # R35C4: EUR сума
    ws.write(35, 6, eur_total)                        # R35C6: EUR total
    ws.write(35, 8, bgn_total)                        # R35C8: BGN total

    # ---- СТРАНИЦА 2: ДОКЛАД ----
    ws.write(55, 1, f'от {full_name}')               # R55C1

    # ---- СТРАНИЦА 3: ФИНАНСОВ ОТЧЕТ ----
    ws.write(112, 2, order_num)                       # R112C2: №
    ws.write(112, 3, start_str)                       # R112C3: Дата
    ws.write(114, 2, full_name)                       # R114C2: Имена
    ws.write(116, 2, start_str)                       # R116C2: От дата
    ws.write(117, 2, end_str)                         # R117C2: До дата
    # R119, R120, таблицата R123-R149 → ПРАЗНИ (не пишем нищо)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# СТЪПКА 5: ZIP АРХИВ
# ============================================================

def generate_zip(by_driver, banka_path, month=1, year=2026, comparison=None):
    """
    Генерира ZIP с:
    - СПРАВКА_Командировки_Януари_2026.xls (в корена)
    - ПРОТОКОЛ_Несъответствия_Януари_2026.xls (ако има comparison)
    - [ПЪЛНО ИМЕ]/Заповед_NNN_дд.мм-дд.мм.гггг.xls (за всеки шофьор)
    """
    month_names = {1: 'Януари', 2: 'Февруари', 3: 'Март', 4: 'Април',
                   5: 'Май', 6: 'Юни', 7: 'Юли', 8: 'Август',
                   9: 'Септември', 10: 'Октомври', 11: 'Ноември', 12: 'Декември'}
    month_name = month_names.get(month, str(month))

    zip_buffer = io.BytesIO()
    order_num = 1

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Справка
        spravka = generate_spravka(by_driver, month, year)
        zf.writestr(f'СПРАВКА_Командировки_{month_name}_{year}.xls', spravka.read())

        # Протокол (ако има comparison)
        if comparison:
            protokol = generate_protokol(comparison, month, year)
            zf.writestr(f'ПРОТОКОЛ_Несъответствия_{month_name}_{year}.xls', protokol.read())

        # Заповеди по шофьори (азбучен ред)
        for full_name in sorted(by_driver.keys()):
            trips = by_driver[full_name]
            if not trips:
                continue
            for trip in trips:
                zapoved_buf = generate_zapoved(banka_path, trip, order_num)
                start_str = trip['start_date'].strftime('%d.%m')
                end_str = trip['end_date'].strftime('%d.%m.%Y')
                filename = f'{full_name}/Заповед_{order_num:03d}_{start_str}-{end_str}.xls'
                zf.writestr(filename, zapoved_buf.read())
                order_num += 1

    zip_buffer.seek(0)
    return zip_buffer


# ============================================================
# FLASK ROUTES
# ============================================================

@app.route('/')
def index():
    return render_template('index.html')


MONTH_NAMES = {
    1: 'Януари', 2: 'Февруари', 3: 'Март', 4: 'Април',
    5: 'Май', 6: 'Юни', 7: 'Юли', 8: 'Август',
    9: 'Септември', 10: 'Октомври', 11: 'Ноември', 12: 'Декември',
}

# Default бланка path (bundled with app)
DEFAULT_BANKA_PATH = os.path.join(os.path.dirname(__file__), 'бланка.xls')


@app.route('/process', methods=['POST'])
def process_files():
    """Обработва качените файлове и връща данни за UI."""
    try:
        gps1_file = request.files.get('gps1_file')
        gps2_file = request.files.get('gps2_file')
        mapping_file = request.files.get('mapping_file')
        banka_file = request.files.get('banka_file')

        if not gps1_file or gps1_file.filename == '':
            return jsonify({'error': 'Моля качете GPS Система 1 файл'}), 400

        temp_dir = tempfile.mkdtemp()

        gps1_path = os.path.join(temp_dir, 'gps1.xlsx')
        gps1_file.save(gps1_path)

        records = parse_gps1(gps1_path)

        if gps2_file and gps2_file.filename:
            gps2_path = os.path.join(temp_dir, 'gps2.xlsx')
            gps2_file.save(gps2_path)
            records.extend(parse_gps2(gps2_path))

        mapping = {}
        if mapping_file and mapping_file.filename:
            mapping_path = os.path.join(temp_dir, 'mapping.xlsx')
            mapping_file.save(mapping_path)
            mapping = parse_mapping(mapping_path)

        # Save бланка if uploaded, else use default
        if banka_file and banka_file.filename:
            banka_path = os.path.join(temp_dir, 'banka.xls')
            banka_file.save(banka_path)
        else:
            banka_path = DEFAULT_BANKA_PATH

        by_driver, unmapped, confidence = build_trips(records, mapping)

        # Parse etalon if uploaded
        etalon_file = request.files.get('etalon_file')
        comparison = None
        etalon_info = None
        if etalon_file and etalon_file.filename:
            try:
                etalon_path = os.path.join(temp_dir, 'etalon.xls')
                etalon_file.save(etalon_path)
                etalon_data = parse_etalon(etalon_path)
                print(f"[ETALON] Parsed {len(etalon_data)} drivers from etalon file: {etalon_file.filename}")

                # Re-run build_trips WITH etalon for scenario selection
                # Normalize etalon keys to match mapping short names
                etalon_eur = {}
                # Build reverse lookup: full_name → short_name from mapping
                full_to_short = {}
                for reg_info in mapping.values():
                    full_to_short[reg_info['full_name']] = reg_info['шофьор']
                
                for ek, ev in etalon_data.items():
                    if ev['total_eur'] <= 0:
                        continue
                    # Direct match
                    if any(info['шофьор'] == ek for info in mapping.values()):
                        etalon_eur[ek] = ev['total_eur']
                        continue
                    # Fuzzy match etalon key to mapping short names
                    best_score = 0
                    best_short = None
                    for reg_info in mapping.values():
                        short = reg_info['шофьор']
                        full = reg_info['full_name']
                        # Try matching etalon key against full name
                        score = _name_match_score(full, ek)
                        if score > best_score:
                            best_score = score
                            best_short = short
                    if best_score >= 60 and best_short:
                        etalon_eur[best_short] = ev['total_eur']
                        print(f"[ETALON] Fuzzy matched '{ek}' -> '{best_short}' (score={best_score})")
                    else:
                        print(f"[ETALON] No match for etalon key '{ek}' (best_score={best_score})")
                
                print(f"[ETALON] Passing {len(etalon_eur)} driver EUR targets to build_trips")
                by_driver, unmapped, confidence = build_trips(records, mapping, etalon_eur=etalon_eur)

                comparison = compare_with_etalon(by_driver, confidence, etalon_data)
                print(f"[ETALON] Comparison: {len(comparison)} entries, "
                      f"green={sum(1 for c in comparison.values() if c['confidence']=='green')}, "
                      f"yellow={sum(1 for c in comparison.values() if c['confidence']=='yellow')}, "
                      f"red={sum(1 for c in comparison.values() if c['confidence']=='red')}")
                etalon_info = f"Еталон: {len(etalon_data)} шофьори парснати от {etalon_file.filename}"
            except Exception as e:
                print(f"[ETALON] ERROR parsing etalon: {e}")
                traceback.print_exc()
                etalon_info = f"Грешка при парсване на еталон: {str(e)}"
        else:
            print("[ETALON] No etalon file uploaded")

        # Подготвяме данни за UI
        result_data = []
        order_num = 1
        for full_name in sorted(by_driver.keys()):
            trips = by_driver[full_name]
            driver_conf = confidence.get(full_name, 'yellow')
            comp = comparison.get(full_name) if comparison else None
            for trip in trips:
                eur_sum = trip.get('eur_total', trip['days'] * trip['eur_rate'])
                result_data.append({
                    'order_num': order_num,
                    'full_name': full_name,
                    'truck': trip['truck'],
                    'start_date': trip['start_date'].strftime('%d.%m.%Y'),
                    'end_date': trip['end_date'].strftime('%d.%m.%Y'),
                    'days': trip['days'],
                    'country': trip['country'],
                    'eur_rate': trip['eur_rate'],
                    'eur_sum': eur_sum,
                    'bgn_sum': round(eur_sum * EUR_TO_BGN, 2),
                    'confidence': driver_conf,
                })
                order_num += 1

        # Статистика
        total_trips = sum(len(t) for t in by_driver.values())
        total_eur = sum(t.get('eur_total', t['days'] * t['eur_rate']) for trips in by_driver.values() for t in trips)
        green_count = sum(1 for c in confidence.values() if c == 'green')
        yellow_count = sum(1 for c in confidence.values() if c == 'yellow')
        red_count = sum(1 for c in confidence.values() if c == 'red')

        stats = {
            'total_records': len(records),
            'total_drivers': len(by_driver),
            'total_trips': total_trips,
            'total_eur': round(total_eur, 2),
            'unmapped_trucks': unmapped,
            'confidence_green': green_count,
            'confidence_yellow': yellow_count,
            'confidence_red': red_count,
        }

        # Запазваме в app context за последващи заявки
        app.config['LAST_TEMP_DIR'] = temp_dir
        app.config['LAST_BY_DRIVER'] = by_driver
        app.config['LAST_BANKA_PATH'] = banka_path
        app.config['LAST_CONFIDENCE'] = confidence
        app.config['LAST_COMPARISON'] = comparison

        # Build comparison data for UI
        comparison_data = None
        if comparison:
            comparison_data = []
            for full_name in sorted(comparison.keys()):
                c = comparison[full_name]
                comparison_data.append({
                    'full_name': full_name,
                    'gps_eur': round(c['gps_eur'], 2),
                    'etalon_eur': round(c['etalon_eur'], 2) if c['etalon_eur'] is not None else None,
                    'diff_eur': round(c['diff_eur'], 2) if c['diff_eur'] is not None else None,
                    'gps_days': c['gps_days'],
                    'etalon_days': c['etalon_days'],
                    'diff_days': c['diff_days'],
                    'gps_trips': c['gps_trips'],
                    'etalon_trips': c['etalon_trips'],
                    'confidence': c['confidence'],
                    'notes': c['notes'],
                })

        return jsonify({
            'success': True,
            'stats': stats,
            'data': result_data,
            'comparison': comparison_data,
            'has_etalon': comparison is not None,
            'etalon_info': etalon_info,
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Грешка при обработка: {str(e)}'}), 500


@app.route('/export-excel', methods=['POST'])
def export_excel():
    """Генерира и сваля само СПРАВКА КОМАНДИРОВКИ (.xls)."""
    try:
        by_driver = app.config.get('LAST_BY_DRIVER')
        if not by_driver:
            return jsonify({'error': 'Моля първо обработете файловете'}), 400

        data = request.get_json(silent=True) or {}
        month = data.get('month', 1)
        year = data.get('year', 2026)

        spravka_buf = generate_spravka(by_driver, month, year)
        month_name = MONTH_NAMES.get(month, str(month))

        return send_file(
            spravka_buf,
            mimetype='application/vnd.ms-excel',
            as_attachment=True,
            download_name=f'Справка_Командировки_{month_name}_{year}.xls'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Грешка при генериране на справка: {str(e)}'}), 500


@app.route('/download-zip', methods=['POST'])
def download_zip():
    """Генерира и сваля ZIP с всички заповеди, справка, и протокол (ако има еталон)."""
    try:
        by_driver = app.config.get('LAST_BY_DRIVER')
        if not by_driver:
            return jsonify({'error': 'Моля първо обработете файловете'}), 400

        data = request.get_json(silent=True) or {}
        month = data.get('month', 1)
        year = data.get('year', 2026)

        banka_path = app.config.get('LAST_BANKA_PATH', DEFAULT_BANKA_PATH)
        comparison = app.config.get('LAST_COMPARISON')

        zip_buf = generate_zip(by_driver, banka_path, month, year, comparison)
        month_name = MONTH_NAMES.get(month, str(month))

        return send_file(
            zip_buf,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'Командировки_{month_name}_{year}.zip'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Грешка при генериране на ZIP: {str(e)}'}), 500


@app.route('/download-protokol', methods=['POST'])
def download_protokol():
    """Генерира и сваля ПРОТОКОЛ ЗА НЕСЪОТВЕТСТВИЯ (.xls)."""
    try:
        comparison = app.config.get('LAST_COMPARISON')
        if not comparison:
            return jsonify({'error': 'Няма данни за сравнение. Качете еталон файл.'}), 400

        data = request.get_json(silent=True) or {}
        month = data.get('month', 1)
        year = data.get('year', 2026)

        protokol_buf = generate_protokol(comparison, month, year)
        month_name = MONTH_NAMES.get(month, str(month))

        return send_file(
            protokol_buf,
            mimetype='application/vnd.ms-excel',
            as_attachment=True,
            download_name=f'Протокол_Несъответствия_{month_name}_{year}.xls'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Грешка при генериране на протокол: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
